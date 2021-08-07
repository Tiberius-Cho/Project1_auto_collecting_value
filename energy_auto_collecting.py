from openpyxl import load_workbook
from openpyxl import Workbook
wb = Workbook()
wb2 = load_workbook("Energydata.xlsx")
ws = wb.active

# 행에 값 입력

ws["B1"] = "메인전기에너지[kwh]"
ws["C1"] = "스키장전기에너지[kwh]"
ws["D1"] = "메인가스에너지[m^2]"
ws["E1"] = "스키장가스에너지[m^2]"
ws["F1"] = "메인상수도[m^2]"
ws["G1"] = "스키장상수도[m^2]"    

sheetlist = wb2.sheetnames # 모든 Sheet 이름 추가
print(sheetlist)
print(sheetlist[0])
print(len(sheetlist)) # 총 시트 수 = 날짜 수

# 열에 날짜 입력

for i in range(0,len(sheetlist)):
    j = i+2
    k = str(j)
    cellname = "A" + k
    ws[cellname] = sheetlist[i]

row_name = 2
for i in range(0,len(sheetlist)):
    new_ws = wb2[sheetlist[i]]

# 전기에너지
    value1 = 0 # 메인 전기에너지
    for x in range(4, 26):
        energyvalue1 = new_ws.cell(row=x, column=4).value
        #print(type(energyvalue1))
        if type(energyvalue1) == int:
            value1 += energyvalue1
        elif type(energyvalue1) == float:
            value1 += energyvalue1
        else:
            pass
    print(sheetlist[i], "메인 전기에너지는", value1)

    value2 = 0 # 스키장 전기에너지
    for x in range(36, 70):
        energyvalue2 = new_ws.cell(row=x, column=4).value
        #print(type(energyvalue2))
        if type(energyvalue2) == int:
            value2 += energyvalue2
        elif type(energyvalue2) == float:
            value2 += energyvalue2
        else:
            pass
    print(sheetlist[i], "스키장 전기에너지는", value2)

# 가스에너지
    value3 = 0 # 메인 가스에너지
    for x in range(4, 26):
        energyvalue3 = new_ws.cell(row=x, column=8).value
        #print(type(energyvalue3))
        if type(energyvalue3) == int:
            value3 += energyvalue3
        elif type(energyvalue3) == float:
            value3 += energyvalue3
        else:
            pass
    print(sheetlist[i], "메인 가스에너지는", value3)

    value4 = 0 # 스키장 가스에너지
    for x in range(36, 70):
        energyvalue4 = new_ws.cell(row=x, column=8).value
        #print(type(energyvalue4))
        if type(energyvalue4) == int:
            value4 += energyvalue4
        elif type(energyvalue4) == float:
            value4 += energyvalue4
        else:
            pass
    print(sheetlist[i], "스키장 가스에너지는", value4)

# 상수도
    value5 = 0 # 메인 상수도
    for x in range(4, 26):
        energyvalue5 = new_ws.cell(row=x, column=12).value
        if type(energyvalue5) == int:
            value5 += energyvalue5
        elif type(energyvalue5) == float:
            value5 += energyvalue5
        else:
            pass
    print(sheetlist[i], "메인 상수도는", value5)

    value6 = 0 # 스키장 상수도
    for x in range(36, 70):
        energyvalue6 = new_ws.cell(row=x, column=12).value
        if type(energyvalue6) == int:
            value6 += energyvalue6
        elif type(energyvalue6) == float:
            value6 += energyvalue6
        else:
            pass
    print(sheetlist[i], "스키장 상수도는", value6)

    # 각 에너지 값을 엑셀에 입력
    ws.cell(column=2, row=row_name, value=value1)
    ws.cell(column=3, row=row_name, value=value2)
    ws.cell(column=4, row=row_name, value=value3)
    ws.cell(column=5, row=row_name, value=value4)
    ws.cell(column=6, row=row_name, value=value5)
    ws.cell(column=7, row=row_name, value=value6)
    row_name += 1

wb.save("Energy_collected.xlsx")